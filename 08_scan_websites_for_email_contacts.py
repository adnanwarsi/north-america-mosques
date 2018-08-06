from bs4 import BeautifulSoup
import re
import json
import sys
import requests
import requests.exceptions
from urllib.parse import urlsplit
from collections import deque
import openpyxl
import csv

def print_usage():
    print ('\nusage1: \n\n>>>python ' + sys.argv[0] + '  {excel file} {(optional) start_row}\n\n')

if len(sys.argv) < 2:
    print_usage()
    exit()

excel_db_file = sys.argv[1]

start_row = 2
if (sys.argv[2]):
    print('apparetnly there is a start row number for source excel file')
    if isinstance(int(sys.argv[2]), int):
        start_row = int(sys.argv[2])
    else:
        print ('is not an integer', sys.argv[2])
        print_usage()
        exit()

print('Will start scanning websites from row number %d', start_row)

wb = openpyxl.load_workbook(excel_db_file)
sheet = wb.get_sheet_by_name('Sheet1')

csvfile = 'emails.csv'

# init for the URL scans
headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'
}


for rowNum in range(start_row, sheet.max_row):  # skip the first row
    masjid_url = sheet.cell(row=rowNum, column=6).value
    # print url
    print (str("\n%d>>>>    will scan the website %s" % (rowNum, masjid_url)))

    # extract base url and path to resolve relative links
    masjid_url_parts = urlsplit(masjid_url)
    masjid_url_base_url = "{0.scheme}://{0.netloc}".format(masjid_url_parts)

    if masjid_url:
        # a queue of urls to be crawled
        new_urls = deque([masjid_url])
        # a set of urls that we have already crawled
        processed_urls = set()
        # a set of crawled emails
        emails = set()
        # process urls one by one until we exhaust the queue
        # ignore the masjid_url that generates too many (more than 20 links); perhaps convoluted website
        while (len(new_urls) < 20) and (len(new_urls) > 0) :
            # move next url from the queue to the set of processed urls
            url = new_urls.popleft()
            processed_urls.add(url)
            # extract base url and path to resolve relative links
            parts = urlsplit(url)
            base_url = "{0.scheme}://{0.netloc}".format(parts)

            if (masjid_url_base_url != base_url):
                # print("ignore %s" % url)
                continue

            # print("base url %s" % base_url)

            path = url[:url.rfind('/') + 1] if '/' in parts.path else url
            # get url's content
            print('.', end='', flush = True)
            # print("Processing %s" % url)
            try:
                response = requests.get(url)

            except (requests.exceptions.MissingSchema, requests.exceptions.ConnectionError):
                # ignore pages with errors
                continue

            # extract all email addresses and add them into the resulting set
            new_emails = set(re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", response.text, re.I))
            # new_emails = set(re.findall(r"[\w\.-]+@[\w\.-]+", response.text, re.I))
            emails.update(new_emails)

            # figure other pages links and extract emails from there
            # create a beutiful soup for the html document
            soup = BeautifulSoup(response.text, "html.parser")
            # find and process all the anchors in the document
            for anchor in soup.find_all("a"):
                # extract link url from the anchor
                link = anchor.attrs["href"] if "href" in anchor.attrs else ''
                # add base url to relative links
                if link.startswith('/'):
                    link = base_url + link
                elif not link.startswith('http'):
                    link = path + link
                # add the new url to the queue if it was not enqueued nor processed yet

                # ignore the links that dont end with a '/'. Probably is a .mp3, .jpg, .pdf etc.
                if link.endswith('/'):
                    # print ('\t\tredirect url %s', link)
                    if not link in new_urls and not link in processed_urls:
                        new_urls.append(link)


        print ('\n'+', '.join(emails))

        # Assuming res is a flat list
        with open(csvfile, "a", newline='') as output:
            writer = csv.writer(output, lineterminator='\n')
            writer.writerow([masjid_url, ', '.join(emails)])
            # for val in [masjid_url, ', '.join(emails)]:
            #    writer.writerow([val])
        output.close()

   
