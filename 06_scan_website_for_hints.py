from bs4 import BeautifulSoup
import re
import json
import sys
import requests
import openpyxl


if len(sys.argv) < 2:
    print ('\n\nusage1: \n\n>>>python ' + sys.argv[0] + '  {excel file} \n\n')
    exit()

excel_db_file = sys.argv[1]

wb = openpyxl.load_workbook(excel_db_file)
sheet = wb.get_sheet_by_name('muslim centers')


# init for the URL scans
headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'
}



for rowNum in range(2, sheet.max_row):  # skip the first row
    if sheet.cell(row=rowNum, column=16).value == 'Sunni':
        url = sheet.cell(row=rowNum, column=9).value
        # print url
        if url:
            print str("%d>>>>    will scan the website %s" % (rowNum, url))
            try:
                r = requests.get(url, headers=headers)
                page = r.text
                bsObj = BeautifulSoup(page, "html.parser")

                if bsObj(text=re.compile('ahmadiyya', re.I)):
                    print 'ahmadiyya msoque found'
                    #sheet.write(rowNum, 16, "Ahmadiyya")
                    sheet.cell(row=rowNum, column=16).value = 'Ahmadiyya'

                elif bsObj(text=re.compile('ismaili', re.I)):
                    print 'Ismaili msoque found'
                    #sheet.write(rowNum, 16, "Ismaili")
                    sheet.cell(row=rowNum, column=16).value = 'Ismaili'

                elif bsObj(text=re.compile('shia |imam ali|bohra |hussain |husain |jafari ', re.I)):
                    print 'shia msoque found'
                    #sheet.write(rowNum, 16, "Shia")
                    sheet.cell(row=rowNum, column=16).value = 'Shia'

                elif bsObj(text=re.compile('nation of islam', re.I)):
                    print 'Nation Of Islam msoque found'
                    #sheet.write(rowNum, 16, "Nation Of Islam")
                    sheet.cell(row=rowNum, column=16).value = 'Nation Of Islam'

                elif bsObj(text=re.compile('sufi ', re.I)):
                    print 'Sufi msoque found'
                    #sheet.write(rowNum, 16, "Sufi")
                    sheet.cell(row=rowNum, column=16).value = 'Sufi'

            except Exception as e:
                print (e)

wb.save(excel_db_file)
