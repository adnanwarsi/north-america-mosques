import sys
import os
import io
import googlemaps
import openpyxl
import time
import json

start_time = time.time()

if len(sys.argv) < 4:
    print ('\n\nusage1: \n\n>>>python ' + sys.argv[0] + '  {excel file with US coordinates} {search term}  (file with your google api key)\n\n')
    exit()

coord_data_file = sys.argv[1]
search_term = sys.argv[2]

handle=open(sys.argv[3],'r+')
Google_API_Key=handle.read()

# google API reference doc https://github.com/googlemaps/google-maps-services-python
#
gmaps = googlemaps.Client(key=Google_API_Key)



''' make a directory if not already, to save individual receipe files as json '''
places_search_dir_path = "./places_search_data/"
try:
    os.makedirs(places_search_dir_path)
except OSError:
    if not os.path.isdir(places_search_dir_path):
        raise


def google_find_and_write_all_places_for_latlon(lat, lon, srch_radius, search_expression):
    print (str('\n\nPlaces query for lat %s, lon %s ' % (lat, lon)))

    results_log_file_name = places_search_dir_path \
                            + 'radial_search_lat' \
                            + str(lat) \
                            + '_lon' \
                            + str(lon) \
                            + '_radius' \
                            + str(srch_radius) \
                            + '_term_' \
                            + search_expression \
                            + '.json'

    if not os.path.isfile(results_log_file_name):


        # the search has not been done yet.
        # places_result = gmaps.places_radar((str(lat), str(lon)), 50000, keyword=search_term)
        places_result = gmaps.places_nearby(
            location={'lat': lat, 'lng': lon},
            radius=srch_radius,
            keyword=search_expression
        )


        if 'OK' in places_result['status']:
            place_dict = {}  # intialize
            for place_data in places_result['results']:
                try:
                    place_detail_result = gmaps.place(place_data['place_id'], language='en-US')

                    if 'OK' == place_detail_result['status']:
                        print ('>>>' + place_detail_result['result']['name'], end ='')
                        if 'mosque' in place_detail_result['result']['types']:
                            # add to dictionary
                            print (' *********************** is a masjid')
                            place_dict[place_detail_result['result']['place_id']] = place_detail_result
                        elif any(substring in place_detail_result['result']['name'] for substring in ['Muslim', 'Islam', 'Muhammad', 'Mosque', 'Masjid', 'Tawheed', 'Hijra', 'Imam ', 'Quran', 'An-Noor', '-Uloom', 'Ahmadiyya', 'Bani Hashim', 'Ismaili', 'Jamatkhana', 'Hussain', 'Al-iman', 'Alhidayat', 'Madinah']):
                            # add to dictionary
                            print (' *********************** is a masjid')
                            place_dict[place_detail_result['result']['place_id']] = place_detail_result
                        else:
                            print ('IGNORE ::: not a masjid')
                    else:
                        print ('IGNORE : Did not get data back for ' + place_data['place_id'])

                except:
                    print ('ERROR in processing ' + place_data['place_id'])

            # write the data into a json file
            with io.open(results_log_file_name, 'w', encoding='utf8') as json_file:
                try:
                    data = json.dumps(place_dict, indent=4, ensure_ascii=False)
                    # unicode(data) auto-decodes data to unicode if str
                    # print (json.dumps(unicode(data), indent=2))
                    # print (data)
                    # json_file.write(unicode(data))
                    json_file.write(data)
                except:
                    # print (json.dumps(unicode(data),indent=2))
                    print ("ERROR::::::::  COULD not write to output file")

    else:
        print (str('IGNORE : %s already exists' % (results_log_file_name)))



# search radius in meters. This is tightly tied with the radial lattice determined for
# search coordinates accross the country
# SEARCHRADIUS = 50000
SEARCHRADIUS = 25000


'''
TEST

Places query for lat 26.412273, lon -80.384851 has several masjids.
Test to see if there are keyword combinations that are more effecient

test_lat = 26.412273
test_lon = -80.384851
google_find_and_write_all_places_for_latlon(test_lat,test_lon,SEARCHRADIUS,search_term)
exit(0)

'''



wb = openpyxl.load_workbook(coord_data_file)
sheet = wb.get_sheet_by_name('Sheet1')


for rowNum in range(2, sheet.max_row):  # skip the first row
    title = sheet.cell(row=rowNum, column=1).value
    lat = sheet.cell(row=rowNum, column=2).value
    lon = sheet.cell(row=rowNum, column=3).value

    print('Node %d of %d'%(rowNum,sheet.max_row))
    google_find_and_write_all_places_for_latlon(lat,lon,SEARCHRADIUS,search_term)



print("--- %s seconds ---" % (time.time() - start_time))
