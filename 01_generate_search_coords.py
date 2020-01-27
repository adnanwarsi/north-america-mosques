import sys
import json
import math
import geopy
from geopy.distance import VincentyDistance
from geopy.geocoders import Nominatim
import openpyxl


geolocator = Nominatim()
def geo_reverse_print(lat, lon):
    print ("\n")
    print(str('%.6f' % lat) + "," + str('%.6f' % lon))
    location = geolocator.reverse(str('%.6f' % lat) + "," + str('%.6f' % lon))
    print(location.address)
    print ("\n")


country = 'no_country'

if len(sys.argv) < 2:
    print ('\n\nusage1: \n\n>>>python ' + sys.argv[0] + ' {argument \'US\' or \'Canada\' exactly} \n\n')
    exit()

country = sys.argv[1]

if (country != 'US' and country != 'Canada'):
    print ("\n\nyou gave the country name not supported: " + country)
    print ('\n\nusage1: \n\n>>>python ' + sys.argv[0] + ' {argument \'US\' or \'Canada\' exactly} \n\n')
    exit()


# given: lat1, lon1, b = bearing in degrees, d = distance in kilometers

# google API reference doc https://github.com/googlemaps/google-maps-services-python
#
# import googlemaps
# from datetime import datetime
#
# gmaps = googlemaps.Client(key='Add Your Key here')
#
# # Geocoding an address
# geocode_result = gmaps.geocode('1600 Amphitheatre Parkway, Mountain View, CA')
#
# # Look up an address with reverse geocoding
# reverse_geocode_result = gmaps.reverse_geocode((40.714224, -73.961452))
#

'''
default to US
'''

# US mainland (excluding Alaska, Hawaii, PeurtoRico) enclosing rectangle
lat_sw_corner = 25.194324
lon_sw_corner = -124.683966
lat_ne_corner = 49.431926
lon_ne_corner = -65.918096

# Canada
if (country == 'Canada'):
    lat_sw_corner = 41.698099
    lon_sw_corner = -141.285412
    lat_ne_corner = 69.921785
    lon_ne_corner = -52.178718


'''
The distance between the centers of the circles in the X-direction is sqrt(3)*R  and not 2*R
because the search radius of google api allows R = 50km
'''
# distance in kilometers
radius = 25
d = math.sqrt(3)*radius
# d = math.sqrt(3)*50  # distance in kilometers


# initialize
lat = lat_sw_corner
lon = lon_sw_corner
row = 0
coord_matrix = {}

while (lat < lat_ne_corner):
    row += 1
    col = 1
    coord_matrix[str('row%02dcol%02d' % (row, col))] = [str('%.6f' % lat) , str('%.6f' % lon)]
    row_origin = geopy.Point(lat, lon)
    while (lon < lon_ne_corner):
        col += 1
        b = 90 # bearing in degrees from north
        origin = geopy.Point(lat, lon)
        destination = VincentyDistance(kilometers=d).destination(origin, b)
        lat, lon = destination.latitude, destination.longitude
        coord_matrix[str('row%02dcol%02d' % (row, col))] = [str('%.6f' % lat) , str('%.6f' % lon)]
        print (str('%s,%s'%(lat,lon)))

    # coord_matrix.append(row_array)
    b = 30*math.pow((-1),(row % 2))  # bearing in degrees from north
    destination = VincentyDistance(kilometers=d).destination(row_origin, b)
    lat, lon = destination.latitude, destination.longitude

# print json.dumps(coord_matrix,indent=2,sort_keys=True)

# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute
sheet = wb.active

# Cell objects also have row, column
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or column integer
# is 1, not 0. Cell object is created by
# using sheet object's cell() method.

# put in the column titles
c1 = sheet.cell(row = 1, column = 1)
c2 = sheet.cell(row = 1, column = 2)
c3 = sheet.cell(row = 1, column = 3)

# writing values to cells
c1.value = "title"
c2.value = "latitude"
c3.value = "longitude"

# start filling in from the second row onwards
rownum = 2
for elem in sorted(coord_matrix):
    c1 = sheet.cell(row=rownum, column=1)
    c2 = sheet.cell(row=rownum, column=2)
    c3 = sheet.cell(row=rownum, column=3)

    c1.value  = elem
    c2.value  = coord_matrix[elem][0]
    c3.value  = coord_matrix[elem][1]
    rownum += 1

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
wb.save(str('%s_search_coords_radius_%d.xlsx')%(country,radius))

# geo_reverse_print(lat2,lon2)

