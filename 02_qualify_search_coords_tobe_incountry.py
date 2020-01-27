import sys
import json
import googlemaps

import openpyxl

if len(sys.argv) < 4:
    print ('\n\nusage1: \n\n>>>python ' + sys.argv[0] + '  {excel file with coordinates} {counrty code} (file with your google api key)\n\n')
    exit()

coord_data_file = sys.argv[1]
country_code = sys.argv[2]

handle=open(sys.argv[3],'r+')
Google_API_Key=handle.read()

# google API reference doc https://github.com/googlemaps/google-maps-services-python
#
gmaps = googlemaps.Client(key=Google_API_Key)

wb = openpyxl.load_workbook(coord_data_file)
sheet = wb.get_sheet_by_name('Sheet')

for rowNum in range(2, sheet.max_row):  # skip the first row
    status = sheet.cell(row=rowNum, column=4).value
    print(str('row %d of %d' % (rowNum, sheet.max_row)) , end = '')
    if status != 'checked':
        # get the status
        title = sheet.cell(row=rowNum, column=1).value
        lat = sheet.cell(row=rowNum, column=2).value
        lon = sheet.cell(row=rowNum, column=3).value
        print (str(' for %s at lat %s, lon %s ' % (title, lat, lon)))
        reverse_geocode_result = gmaps.reverse_geocode((str(lat), str(lon)))
        # print (json.dumps(reverse_geocode_result,indent=2))
        try:
            if 'address_components' in reverse_geocode_result[0]:
                for elem in reverse_geocode_result[0]['address_components']:
                    # print ('>>>>>>>> %%%%%%%%%%%%%' +  json.dumps(elem,indent=2))
                    if 'country' == elem['types'][0]:
                        print (elem['short_name'])
                        sheet.cell(row=rowNum, column=5).value = elem['short_name']
                        if country_code == elem['short_name']:
                            print ('***** coordinates are in ' + country_code)

        except:
            print ('This coordinate is not in the' + country_code)

        sheet.cell(row=rowNum, column=4).value = 'checked'

wb.save(coord_data_file)
