import requests
import json
import openpyxl
# import codecs

import sys
reload(sys)
sys.setdefaultencoding('utf-8')


if len(sys.argv) < 2:
    print ('\n\nusage1: \n\n>>>python ' + sys.argv[0] + '  {excel file} \n\n')
    exit()

excel_db_file = sys.argv[1]

wb = openpyxl.load_workbook(excel_db_file)
sheet = wb.get_sheet_by_name('muslim centers')


sheet.cell(row=1, column=17).value = 'FSQ Name' # column Q
sheet.cell(row=1, column=18).value = 'FSQ Users Count' # column R

# Foursquare parameters
client_id = '2FSOIJFEHCLIHYB25WRI20YHSPEYJ2HMLMR5AUSSE4AO5MGF'
secret_id = 'MDSKWDSSIHSLIWBYRF2L5RDZEEYJYWWRWAZ5ZJQLPSETZRVW'


for rowNum in range(2, sheet.max_row):  # skip the first row
    # print (rowNum)
    # query = unicode(sheet.cell(row=rowNum, column=4).value , 'utf8')# name of the masjid
    query = sheet.cell(row=rowNum, column=4).value # name of the masjid
    lat = sheet.cell(row=rowNum, column=2).value # lat of the masjid
    lon = sheet.cell(row=rowNum, column=3).value # lon of the masjid
    radius = 100

    url = str("https://api.foursquare.com/v2/venues/search?v=20170101&client_id=%s&client_secret=%s&ll=%s,%s&intent=browse&radius=%d&query=\"%s\"" % (client_id, secret_id, lat, lon, radius, query))
    # print (url)

    print('\nrow %d : %s'%(rowNum,query))

    req = requests.request('GET', url)
    if  req.status_code == 200:
        resp = req.json()
        # print(json.dumps(resp, indent=2))

        if len(resp['response']['venues'])>0:
            fsq_name = resp['response']['venues'][0]['name']
            fsq_stats_users_count = resp['response']['venues'][0]['stats']['usersCount']

            print('foursquare name is %s with users count %d'%(fsq_name,fsq_stats_users_count))

            sheet.cell(row=rowNum, column=17).value = fsq_name  # column Q
            sheet.cell(row=rowNum, column=18).value = fsq_stats_users_count  # column R

        else:
            print('did not get FourSquare info')


wb.save(excel_db_file)
